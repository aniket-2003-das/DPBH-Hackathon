from flask import Flask, jsonify, request, send_file

from flask_cors import CORS
from joblib import load
import pandas as pd
import base64
import openpyxl
from openpyxl.drawing.image import Image
from io import BytesIO
from zipfile import ZipFile
import json
import numpy as np

presence_classifier = load('presence_classifer.joblib')
presence_vect = load('presence_vectorizer.joblib')
category_classifier = load('category_classifier.joblib')
category_vect = load('category_vectorizer.joblib')

app = Flask(__name__)
CORS(app)


@app.route('/', methods=['POST'])
def main():
    if request.method == 'POST':
        output = []
        data = request.get_json().get('tokens')

        for token in data:
            result = presence_classifier.predict(
                presence_vect.transform([token]))
            if result == 'Dark':
                cat = category_classifier.predict(
                    category_vect.transform([token]))
                output.append(cat[0])
            else:
                output.append(result[0])

        dark = [data[i] for i in range(len(output)) if output[i] == 'Dark']
        for d in dark:
            print(d)
        print()
        print(len(dark))

        message = '{ \'result\': ' + str(output) + ' }'

        json = jsonify(message)

        return json


@app.route('/report', methods=['POST'])
def report():
    if request.method == 'POST':
        data = request.get_json()
        image_data = data.get('image')
        description = data.get('description')

        image_bytes = base64.b64decode(image_data.split(',')[1])

        try:
            workbook = openpyxl.load_workbook('reports.xlsx')
            csv_file_path = 'path/to/your/dataset.csv'
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        worksheet = workbook.active
        if worksheet.title != 'Reports':
            worksheet = workbook.create_sheet(title='Reports')

        
        worksheet.append([description])

        image_stream = BytesIO(image_bytes)

        img = Image(image_stream)
        worksheet.add_image(img, f'B{worksheet.max_row + 1}')

        workbook.save('reports.xlsx')

        return jsonify({'message': 'Report saved successfully.'})





if __name__ == '__main__':
    app.run(threaded=True, debug=True)
